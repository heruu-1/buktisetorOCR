## backend/bukti_setor/services/excel_exporter_bukti_setor.py
# # ==============================================================================

import os
import tempfile
from flask import send_file, jsonify
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, Side
# ==============================================================================
# File: backend/bukti_setor/services/excel_exporter_bukti_setor.py

def generate_excel_bukti_setor_export(db):
    try:
        from models import BuktiSetor  # Pastikan model ini diimpor dari tempat yang benar

        BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
        template_path = os.path.join(BASE_DIR, "templates", "rekap_template_bukti_setor.xlsx")
        template_path = os.path.normpath(template_path)
        
        print(f"[DEBUG] Path template bukti setor: {template_path}")
        print(f"[DEBUG] Exists? {os.path.exists(template_path)}")
        # Cek apakah template ada
        if not os.path.exists(template_path):
            return jsonify(error=f"Template tidak ditemukan di path: {template_path}"), 404

        wb = load_workbook(template_path)
        ws = wb.active

        # Ambil data dari database
        data = db.session.execute(
            db.select(BuktiSetor).order_by(BuktiSetor.tanggal.desc())
        ).scalars().all()

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        start_row = 2  # Anggap template punya judul di baris 1-2
        for idx, item in enumerate(data, start=start_row):
            ws.cell(row=idx, column=1, value=item.tanggal.strftime("%Y-%m-%d")).border = thin_border
            ws.cell(row=idx, column=2, value=item.kode_setor).border = thin_border

            jumlah_cell = ws.cell(row=idx, column=3, value=float(item.jumlah))
            jumlah_cell.number_format = "#,##0.00"
            jumlah_cell.border = thin_border

            ws.cell(row=idx, column=4, value=item.created_at.strftime("%Y-%m-%d %H:%M:%S")).border = thin_border

        # Simpan ke file sementara
        temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp.name)
        print(f"[✅] Export Bukti Setor berhasil ke: {temp.name}")

        return send_file(
            temp.name,
            as_attachment=True,
            download_name="rekap_bukti_setor.xlsx"
        )

    except Exception as e:
        print(f"[❌] Error generate Excel Bukti Setor: {e}")
        return jsonify({"error": str(e)}), 500

